"""
FastAPI backend for the Account Reconciliation prototype.

Run:
    pip install -r requirements.txt
    uvicorn main:app --reload --port 8000
"""
import base64
import csv
import io
import json
import os
import re as _re
import uuid
from datetime import datetime, timedelta
from calendar import monthrange
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from db import init_db, seed_users, tx, log


# ────────────────────────── .env loader (no external dep) ──────────────────────────
# Populates os.environ from these files in order — first hit wins per key, and
# values already in the real shell env always win over file values. Lets users
# drop their ANTHROPIC_API_KEY into a file instead of having to remember to
# export it in every shell that runs uvicorn.
def _load_env_files():
    candidates = [
        Path(__file__).parent / ".env",
        Path(__file__).parent.parent / ".env",
        Path.home() / ".recon-app" / ".env",
    ]
    for p in candidates:
        if not p.exists():
            continue
        try:
            for raw in p.read_text().splitlines():
                line = raw.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                # Don't clobber values that are actually set in the real shell
                # env — but treat an empty string as "not set" so a blank value
                # left over in the process env doesn't block the .env file.
                if k and not os.environ.get(k):
                    os.environ[k] = v
            print(f"[env] loaded {p}")
        except Exception as e:
            print(f"[env] could not read {p}: {e}")


_load_env_files()

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from claude_agent_sdk import (
        query as _sdk_query,
        ClaudeAgentOptions,
        CLINotFoundError,
        CLIConnectionError,
        AssistantMessage,
        ResultMessage,
        TextBlock,
    )
    HAS_AGENT_SDK = True
except ImportError:
    HAS_AGENT_SDK = False


def _find_claude_cli() -> Optional[str]:
    """Locate the `claude` CLI binary so the Agent SDK subprocess can launch it
    regardless of what's on the shell PATH (uvicorn often runs with a minimal
    PATH that excludes ~/.local/bin)."""
    import shutil
    hit = shutil.which("claude")
    if hit:
        return hit
    for p in (
        Path.home() / ".local" / "bin" / "claude",
        Path.home() / ".claude" / "local" / "claude",
        Path("/opt/homebrew/bin/claude"),
        Path("/usr/local/bin/claude"),
    ):
        if p.exists():
            return str(p)
    return None

# ────────────────────────── App setup ──────────────────────────
app = FastAPI(title="Osfin - Account Reconciliation")

# CORS configuration - allow configured origins in production, all in dev
_cors_origins = os.environ.get("CORS_ORIGINS", "http://localhost:3000,http://localhost:5173").split(",")
_cors_origins = [o.strip() for o in _cors_origins]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins if os.environ.get("CORS_ORIGINS") else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path(os.environ.get("RECON_UPLOAD_DIR", str(Path(__file__).parent / "uploads")))
UPLOAD_DIR.mkdir(exist_ok=True, parents=True)


@app.get("/health")
def health():
    """Health check endpoint for monitoring/load balancers."""
    return {"status": "ok"}


@app.on_event("startup")
def startup():
    init_db()
    seed_users()


# ────────────────────────── Auth ──────────────────────────
class LoginReq(BaseModel):
    username: str
    password: str


def _user_matches(assigned: Optional[str], user: dict) -> bool:
    """A reconciliation is assigned to `assigned` — match against either username or name."""
    if not assigned:
        return False
    a = str(assigned).strip().lower()
    return a == user["username"].lower() or a == user["name"].lower()


def _visible_to(row: dict, user: dict, role_field: str) -> bool:
    """A reconciliation is visible to a Preparer/Approver if it's either assigned
    to them, or unassigned (part of the pool). This lets a fresh import with no
    Preparer/Approver columns still be picked up for work."""
    assigned = row.get(role_field)
    if not assigned:
        return True
    return _user_matches(assigned, user)


def get_user(authorization: Optional[str] = Header(None), token: Optional[str] = None):
    """Token = username. Accepts either:
       • Authorization header: 'Bearer <username>' (preferred, used by fetch calls)
       • ?token= query string (used for <a href> downloads that can't set headers)
    """
    username = None
    if authorization and authorization.startswith("Bearer "):
        username = authorization.replace("Bearer ", "").strip()
    elif token:
        username = token.strip()
    if not username:
        raise HTTPException(status_code=401, detail="Not authenticated")
    with tx() as conn:
        row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if not row:
            raise HTTPException(status_code=401, detail="Invalid token")
        return dict(row)


def require_role(*allowed):
    def _dep(user=Depends(get_user)):
        if user["role"] not in allowed:
            raise HTTPException(status_code=403, detail=f"Requires role: {', '.join(allowed)}")
        return user
    return _dep


@app.post("/api/login")
def login(req: LoginReq):
    with tx() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE username = ? AND password = ?",
            (req.username.strip().lower(), req.password),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=401, detail="Invalid username or password")
        u = dict(row)
        u.pop("password", None)
        log(req.username, "login")
        return {"token": u["username"], "user": u}


@app.get("/api/me")
def me(user=Depends(get_user)):
    u = dict(user)
    u.pop("password", None)
    return u


@app.get("/api/users")
def users(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute("SELECT id, username, name, email, role FROM users").fetchall()
        return [dict(r) for r in rows]


# ────────────────────────── Account Groups ──────────────────────────
class GroupReq(BaseModel):
    name: str
    description: Optional[str] = None


@app.get("/api/groups")
def list_groups(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute("SELECT * FROM account_groups ORDER BY name").fetchall()
        out = []
        for g in rows:
            member_ids = [r["id"] for r in conn.execute(
                "SELECT id FROM accounts WHERE group_id=?", (g["id"],)
            ).fetchall()]
            d = dict(g)
            d["member_count"] = len(member_ids)
            d["member_account_ids"] = member_ids
            out.append(d)
        return out


@app.post("/api/groups")
def create_group(req: GroupReq, user=Depends(require_role("Admin"))):
    gid = f"grp-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        try:
            conn.execute(
                "INSERT INTO account_groups (id, name, description) VALUES (?, ?, ?)",
                (gid, req.name.strip(), req.description),
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))
    log(user["username"], "create_group", "group", gid, {"name": req.name})
    return {"id": gid}


@app.put("/api/groups/{gid}")
def update_group(gid: str, req: GroupReq, user=Depends(require_role("Admin"))):
    with tx() as conn:
        conn.execute(
            "UPDATE account_groups SET name=?, description=? WHERE id=?",
            (req.name.strip(), req.description, gid),
        )
    return {"ok": True}


@app.delete("/api/groups/{gid}")
def delete_group(gid: str, user=Depends(require_role("Admin"))):
    with tx() as conn:
        conn.execute("DELETE FROM account_groups WHERE id=?", (gid,))
    log(user["username"], "delete_group", "group", gid)
    return {"ok": True}


class GroupAssignReq(BaseModel):
    account_ids: List[str]


class RejectReq(BaseModel):
    reason: Optional[str] = None


@app.post("/api/groups/{gid}/members")
def assign_group_members(gid: str, req: GroupAssignReq, user=Depends(require_role("Admin"))):
    """Set group membership to exactly the given account_ids (others in this group
    are cleared)."""
    with tx() as conn:
        g = conn.execute("SELECT * FROM account_groups WHERE id=?", (gid,)).fetchone()
        if not g:
            raise HTTPException(status_code=404, detail="Group not found")
        # Clear existing members
        conn.execute("UPDATE accounts SET group_id=NULL WHERE group_id=?", (gid,))
        # Assign new members
        for aid in req.account_ids:
            conn.execute("UPDATE accounts SET group_id=? WHERE id=?", (gid, aid))
    log(user["username"], "assign_group", "group", gid, {"count": len(req.account_ids)})
    return {"ok": True, "count": len(req.account_ids)}


# ────────────────────────── Group-level certify / approve / reject ──────────────────────────
# A grouped reconciliation is treated as a single reconciliation unit: the
# combined GL balance is what matters, so supporting items may span any member
# recon. These endpoints apply the status transition to every member in one
# shot, checking tolerance against the *combined* totals rather than
# per-member (which can net to zero even if individual members don't).

def _group_members_for_period(conn, gid: str, period: str):
    return conn.execute(
        """SELECT r.*, a.cert_threshold_pct AS _pct, a.cert_threshold_amt AS _amt,
                  a.account_number AS _acct, a.template AS _template
             FROM reconciliations r
             JOIN accounts a ON a.id = r.account_id
            WHERE a.group_id=? AND r.period=?""",
        (gid, period),
    ).fetchall()


@app.post("/api/groups/{gid}/certify")
def group_certify(gid: str, period: str, user=Depends(require_role("Preparer", "Admin"))):
    if not period:
        raise HTTPException(status_code=400, detail="period is required (YYYY-MM)")
    with tx() as conn:
        members = _group_members_for_period(conn, gid, period)
        if not members:
            raise HTTPException(status_code=404, detail="No reconciliations for this group in this period")

        combined_gl = 0.0
        combined_items = 0.0
        # Combined tolerance: sum per-account amount thresholds; take the max
        # pct threshold as a percentage of combined GL.
        max_pct = 0.0
        sum_amt = 0.0
        for m in members:
            combined_gl += float(m["gl_balance"] or 0)
            member_items = conn.execute(
                "SELECT * FROM supporting_items WHERE recon_id=?", (m["id"],),
            ).fetchall()
            combined_items += _effective_items_sum(member_items, m["_template"], m["period"])
            max_pct = max(max_pct, float(m["_pct"] or 0))
            sum_amt += float(m["_amt"] or 0)

        diff = combined_gl - combined_items
        tolerance = max(sum_amt, abs(combined_gl) * max_pct / 100.0, 0.01)
        if abs(diff) > tolerance:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot certify group. Combined unidentified difference "
                       f"{diff:.2f} exceeds tolerance {tolerance:.2f}",
            )

        transitioned = 0
        skipped = 0
        ts = datetime.now().strftime("%m/%d/%Y")
        for m in members:
            # Skip anything already at or past Pending Approval — certify is
            # only meaningful from Not Prepared / In Progress / Rejected.
            if m["status"] in ("Pending Approval", "Reviewed", "Approved", "System Certified"):
                skipped += 1
                continue
            conn.execute(
                """UPDATE reconciliations SET status='Pending Approval', prep_date=?,
                      certified_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (ts, user["name"], m["id"]),
            )
            transitioned += 1
    log(user["username"], "group_certify", "group", gid,
        {"period": period, "members": len(members), "transitioned": transitioned})
    return {"ok": True, "group_id": gid, "period": period,
            "transitioned": transitioned, "skipped": skipped,
            "combined_unidentified": round(diff, 2)}


@app.post("/api/groups/{gid}/approve")
def group_approve(gid: str, period: str, user=Depends(require_role("Approver", "Admin"))):
    if not period:
        raise HTTPException(status_code=400, detail="period is required (YYYY-MM)")
    transitioned = 0
    ts = datetime.now().strftime("%m/%d/%Y")
    with tx() as conn:
        members = _group_members_for_period(conn, gid, period)
        if not members:
            raise HTTPException(status_code=404, detail="No reconciliations for this group in this period")
        for m in members:
            if m["status"] != "Pending Approval":
                continue
            conn.execute(
                """UPDATE reconciliations SET status='Reviewed', app_date=?,
                      approved_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (ts, user["name"], m["id"]),
            )
            transitioned += 1
    log(user["username"], "group_approve", "group", gid,
        {"period": period, "transitioned": transitioned})
    return {"ok": True, "transitioned": transitioned}


@app.post("/api/groups/{gid}/reject")
def group_reject(gid: str, period: str, req: RejectReq,
                 user=Depends(require_role("Approver", "Admin"))):
    if not period:
        raise HTTPException(status_code=400, detail="period is required (YYYY-MM)")
    transitioned = 0
    with tx() as conn:
        members = _group_members_for_period(conn, gid, period)
        if not members:
            raise HTTPException(status_code=404, detail="No reconciliations for this group in this period")
        for m in members:
            if m["status"] != "Pending Approval":
                continue
            conn.execute(
                """UPDATE reconciliations SET status='In Progress', reject_reason=?,
                      updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (req.reason, m["id"]),
            )
            if req.reason:
                conn.execute(
                    "INSERT INTO comments (id, recon_id, author, text) VALUES (?, ?, ?, ?)",
                    (f"c-{uuid.uuid4().hex[:8]}", m["id"], user["name"], f"[REJECTED] {req.reason}"),
                )
            transitioned += 1
    log(user["username"], "group_reject", "group", gid,
        {"period": period, "reason": req.reason, "transitioned": transitioned})
    return {"ok": True, "transitioned": transitioned}


# ────────────────────────── Accounts ──────────────────────────
@app.get("/api/accounts")
def list_accounts(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute("SELECT * FROM accounts ORDER BY entity_code, account_number").fetchall()
        return [dict(r) for r in rows]


class AccountReq(BaseModel):
    entity: str
    entity_code: str
    account_number: str
    description: str
    template: str = "General List"
    preparer: Optional[str] = None
    approver: Optional[str] = None
    currency: str = "USD"
    cert_threshold_pct: float = 0
    cert_threshold_amt: float = 0
    group_id: Optional[str] = None


@app.post("/api/accounts")
def create_account(req: AccountReq, user=Depends(require_role("Admin"))):
    aid = f"acc-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        try:
            conn.execute(
                """INSERT INTO accounts
                   (id, entity, entity_code, account_number, description, template,
                    preparer, approver, currency, cert_threshold_pct, cert_threshold_amt, group_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (aid, req.entity, req.entity_code, req.account_number, req.description,
                 req.template, req.preparer, req.approver, req.currency,
                 req.cert_threshold_pct, req.cert_threshold_amt, req.group_id),
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))
    log(user["username"], "create_account", "account", aid, {"account": req.account_number})
    return {"id": aid}


@app.put("/api/accounts/{aid}")
def update_account(aid: str, req: AccountReq, user=Depends(require_role("Admin"))):
    with tx() as conn:
        conn.execute(
            """UPDATE accounts SET entity=?, entity_code=?, account_number=?, description=?,
               template=?, preparer=?, approver=?, currency=?, cert_threshold_pct=?, cert_threshold_amt=?, group_id=?
               WHERE id=?""",
            (req.entity, req.entity_code, req.account_number, req.description, req.template,
             req.preparer, req.approver, req.currency, req.cert_threshold_pct,
             req.cert_threshold_amt, req.group_id, aid),
        )
    return {"ok": True}


@app.delete("/api/accounts/{aid}")
def delete_account(aid: str, user=Depends(require_role("Admin"))):
    with tx() as conn:
        conn.execute("DELETE FROM accounts WHERE id=?", (aid,))
    return {"ok": True}


class TemplatePatchReq(BaseModel):
    template: str


@app.patch("/api/accounts/{aid}/template")
def patch_account_template(aid: str, req: TemplatePatchReq, user=Depends(require_role("Admin"))):
    """Lightweight endpoint for the Admin's inline Template dropdown on the
    Summary grid — avoids having to round-trip the full account payload."""
    if req.template not in ("General List", "Amortizable", "Accrual", "Schedule List"):
        raise HTTPException(status_code=400, detail="Unknown template")
    with tx() as conn:
        hit = conn.execute("SELECT id FROM accounts WHERE id=?", (aid,)).fetchone()
        if not hit:
            raise HTTPException(status_code=404, detail="Account not found")
        conn.execute("UPDATE accounts SET template=? WHERE id=?", (req.template, aid))
    log(user["username"], "change_template", "account", aid, {"template": req.template})
    return {"ok": True, "template": req.template}


# ────────────────────────── Upload (CSV/Excel) ──────────────────────────
def _read_csv(content: bytes):
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_xlsx(content: bytes):
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None or v == "" for v in r):
            continue
        out.append({headers[i]: r[i] for i in range(len(headers))})
    return out


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_").replace("-", "_")


# Accept many column-name variations
FIELD_ALIASES = {
    "entity":         ["entity", "legal_entity", "company", "company_name"],
    "entity_code":    ["entity_code", "entitycode", "entity_cd", "company_code",
                       "company_cd", "co_code", "co", "dept_code"],
    "account_number": ["account", "account_number", "acct", "acct_no", "gl_account",
                       "account_segment", "gl_acct", "gl_account_number"],
    "description":    ["description", "account_description", "acct_desc", "name",
                       "acct_description"],
    "template":       ["template", "template_type", "recon_template"],
    "preparer":       ["preparer", "preparer_name", "prepared_by"],
    "approver":       ["approver", "approver_name", "reviewer", "approved_by"],
    "gl_balance":     ["gl_balance", "balance", "ending_bal", "ending_balance",
                       "net_balance", "net_bal", "closing_balance"],
    "debit_balance":  ["debit_balance", "debit_bal", "debit", "dr"],
    "credit_balance": ["credit_balance", "credit_bal", "credit", "cr"],
    "currency":       ["currency", "ccy"],
    "threshold_pct":  ["threshold_%", "threshold_pct", "threshold_percent", "cert_threshold_pct"],
    "threshold_amt":  ["threshold_amount", "threshold_amt", "cert_threshold_amt"],
    "account_type":   ["account_type", "acct_type", "type"],
    "fiscal_year":    ["fiscal_year", "fy"],
    "fiscal_period":  ["fiscal_period", "period", "period_num"],
    "period_end":     ["period_end_date", "period_end", "as_of_date", "as_of"],
}

def _find(row: dict, field: str, default=None):
    nrow = { _norm(k): v for k, v in row.items() }
    for alias in FIELD_ALIASES.get(field, [field]):
        if alias in nrow and nrow[alias] not in (None, ""):
            return nrow[alias]
    return default


@app.post("/api/upload")
async def upload_tb(
    period: Optional[str] = Form(None),
    classify: bool = Form(False),
    file: UploadFile = File(...),
    user=Depends(require_role("Admin")),
):
    """Upload a trial-balance file (CSV or XLSX). Upserts accounts and creates/updates
    reconciliations for the given period, preserving reconciliation work.

    The `period` form field is optional — if omitted the importer tries to derive it
    from columns like Fiscal_Year / Fiscal_Period or Period_End_Date.

    If `classify=true`, after the import completes, hand every newly-created
    account to the Claude Agent SDK for template classification (General List /
    Amortizable / Accrual / Schedule List). Pre-existing accounts are left
    untouched — the Admin may have set their templates explicitly already.
    """
    content = await file.read()
    result = await _ingest_trial_balance(
        content, file.filename or "upload",
        period=period, classify=classify,
        username=user["username"],
    )
    return result


async def _ingest_trial_balance(content: bytes, filename: str, *,
                                period: Optional[str], classify: bool,
                                username: str) -> dict:
    """Shared ingestion pipeline used by:
       • POST /api/upload             (interactive Admin upload)
       • Scheduler                    (pulls a file from SFTP/S3/local)
    Writes the file to uploads/ for audit, parses rows, upserts accounts and
    reconciliations, optionally runs the Claude classifier for new accounts,
    and returns a structured result dict.
    """
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext == "csv":
        rows = _read_csv(content)
    elif ext in ("xlsx", "xlsm"):
        rows = _read_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")

    if not rows:
        raise HTTPException(status_code=400, detail="File appears to be empty")

    if not period:
        period = _derive_period(rows[0])
    if not period:
        raise HTTPException(
            status_code=400,
            detail="Could not determine period. Either pass a period (YYYY-MM) "
                   "or include Fiscal_Year + Fiscal_Period or Period_End_Date columns."
        )

    import_id = f"imp-{uuid.uuid4().hex[:8]}"
    save_path = UPLOAD_DIR / f"{import_id}__{filename}"
    save_path.write_bytes(content)

    created, updated, reopened = 0, 0, 0
    skipped = 0
    errors: List[str] = []
    skip_reasons: dict = {}
    new_accounts: List[dict] = []  # populated only for rows that CREATE a new account

    def _skip(reason: str, line: int):
        nonlocal skipped
        skipped += 1
        skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
        if len(errors) < 10:
            errors.append(f"Row {line}: {reason}")

    with tx() as conn:
        for i, r in enumerate(rows, start=2):
            try:
                # Skip fully-blank rows silently
                if all(v is None or str(v).strip() == "" for v in r.values()):
                    continue

                entity = str(_find(r, "entity") or "").strip()
                ec     = str(_find(r, "entity_code") or "").strip()
                an     = str(_find(r, "account_number") or "").strip()
                desc   = str(_find(r, "description") or "").strip()
                tpl    = str(_find(r, "template") or "").strip()
                prep   = _find(r, "preparer")
                appr   = _find(r, "approver")
                cur    = str(_find(r, "currency") or "USD").strip()
                acct_type = str(_find(r, "account_type") or "").strip().lower()

                # GL Balance: prefer explicit column, else derive from debit − credit
                bal_raw = _find(r, "gl_balance")
                if bal_raw in (None, ""):
                    dr = _num(_find(r, "debit_balance"))
                    cr = _num(_find(r, "credit_balance"))
                    if dr or cr:
                        bal = dr - cr
                        bal_raw = bal
                    else:
                        bal_raw = 0
                try:
                    bal = float(str(bal_raw).replace(",", "").replace("$", "").strip() or 0)
                except (ValueError, TypeError):
                    bal = 0
                try:
                    tpct = float(str(_find(r, "threshold_pct", 0) or 0).strip())
                except (ValueError, TypeError):
                    tpct = 0
                try:
                    tamt = float(str(_find(r, "threshold_amt", 0) or 0).replace(",", "").replace("$", "").strip() or 0)
                except (ValueError, TypeError):
                    tamt = 0

                if not an:
                    _skip("missing account number", i)
                    continue
                if not ec:
                    ec = entity or "01"
                if not entity:
                    entity = f"Entity {ec}"
                if not desc:
                    desc = f"Account {an}"

                # Default every new account to General List. The caller can
                # pass ?classify=true to re-classify below, or the Admin can
                # change templates individually later.
                if tpl not in ("General List", "Amortizable", "Accrual", "Schedule List"):
                    tpl = "General List"

                existing = conn.execute(
                    "SELECT * FROM accounts WHERE entity_code=? AND account_number=?",
                    (ec, an),
                ).fetchone()
                if existing:
                    account_id = existing["id"]
                    conn.execute(
                        """UPDATE accounts SET description=?, template=?, preparer=?, approver=?,
                           currency=?, cert_threshold_pct=?, cert_threshold_amt=? WHERE id=?""",
                        (desc or existing["description"], tpl,
                         prep or existing["preparer"], appr or existing["approver"],
                         cur or existing["currency"], tpct or existing["cert_threshold_pct"],
                         tamt or existing["cert_threshold_amt"], account_id),
                    )
                    updated += 1
                else:
                    account_id = f"acc-{uuid.uuid4().hex[:8]}"
                    conn.execute(
                        """INSERT INTO accounts (id, entity, entity_code, account_number, description,
                           template, preparer, approver, currency, cert_threshold_pct, cert_threshold_amt)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (account_id, entity, ec, an, desc, tpl, prep, appr, cur, tpct, tamt),
                    )
                    created += 1
                    new_accounts.append({
                        "id": account_id,
                        "account_number": an,
                        "description": desc,
                        "account_type": acct_type,
                    })

                # Reconciliation for this period
                rec = conn.execute(
                    "SELECT * FROM reconciliations WHERE account_id=? AND period=?",
                    (account_id, period),
                ).fetchone()
                if rec:
                    prev_bal = rec["gl_balance"]
                    # PRD: re-import updates balances. If certified rec has changed balance, reopen.
                    new_status = rec["status"]
                    if abs((prev_bal or 0) - bal) > 0.001 and rec["status"] in ("Approved", "Reviewed", "System Certified"):
                        new_status = "In Progress"
                        reopened += 1
                    conn.execute(
                        "UPDATE reconciliations SET gl_balance=?, status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (bal, new_status, rec["id"]),
                    )
                else:
                    rid = f"rec-{uuid.uuid4().hex[:8]}"
                    conn.execute(
                        """INSERT INTO reconciliations (id, account_id, period, gl_balance, status)
                           VALUES (?, ?, ?, ?, 'Not Prepared')""",
                        (rid, account_id, period, bal),
                    )
            except Exception as e:
                if len(errors) < 10:
                    errors.append(f"Row {i}: {e}")
                skipped += 1

        conn.execute(
            """INSERT INTO imports (id, period, filename, uploaded_by, row_count, accounts_created, accounts_updated)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (import_id, period, filename, username, len(rows), created, updated),
        )

    log(username, "upload_tb", "import", import_id,
        {"period": period, "rows": len(rows), "created": created, "updated": updated, "skipped": skipped})

    # Detect "every row was skipped because headers weren't recognised"
    warnings: List[str] = []
    if skipped == len(rows) and len(rows) > 0:
        found_headers = sorted(rows[0].keys()) if rows else []
        warnings.append(
            f"No rows could be imported — column headers weren't recognised. "
            f"Found columns: {', '.join(found_headers)}"
        )

    # Optional: ask Claude to classify each new account's reconciliation template.
    classify_result = None
    if classify and new_accounts:
        try:
            classified = await _classify_templates(new_accounts)
            # Apply template updates in a single transaction
            with tx() as conn:
                for acc_id, tmpl in classified.items():
                    conn.execute("UPDATE accounts SET template=? WHERE id=?", (tmpl, acc_id))
            classify_result = {
                "total":       len(new_accounts),
                "classified":  len(classified),
                "by_template": _tally_templates(classified),
            }
            log(username, "classify_templates", "import", import_id, classify_result)
        except Exception as e:
            warnings.append(f"Template classification failed: {e}. Accounts default to General List.")
            classify_result = {"error": str(e)}

    return {
        "import_id": import_id,
        "period": period,
        "row_count": len(rows),
        "accounts_created": created,
        "accounts_updated": updated,
        "reopened": reopened,
        "skipped": skipped,
        "skip_reasons": skip_reasons,
        "warnings": warnings,
        "errors": errors[:20],
        "classify": classify_result,
    }


def _tally_templates(classified: dict) -> dict:
    """Count how many accounts landed in each template."""
    out = {}
    for t in classified.values():
        out[t] = out.get(t, 0) + 1
    return out


_VALID_TEMPLATES = {"General List", "Amortizable", "Accrual", "Schedule List"}

_CLASSIFY_PROMPT_TMPL = """You are classifying GL accounts for a month-end reconciliation system.
For each account below, choose the BEST-fit reconciliation template from these four options:

- "General List": the default for most balance-sheet accounts. Use this for cash, AR, AP,
  inventory, most revenue/expense accounts — anything where the balance is explained by
  listing the individual items that make it up, without a recurring schedule.
- "Amortizable": prepaid expenses that are paid up-front and expensed over multiple
  periods. Prepaid insurance, prepaid rent, prepaid software licenses, deferred charges,
  intangible asset amortization, bond premium/discount amortization.
- "Accrual": liabilities that grow monthly and are periodically paid out. Accrued bonus,
  accrued compensation, accrued interest, accrued expenses, accrued payroll taxes,
  accrued audit fees WITHOUT a recurring schedule.
- "Schedule List": recurring periodic accruals with irregular settlements tied to a
  specific vendor / counterparty. Audit-fee retainers paid 2-3x/year, legal retainers,
  long-term service contracts with recurring scheduled charges.

Accounts to classify:
{accounts}

Return EXACTLY one JSON object and NOTHING else. No code fence, no prose:
{{"classifications": [{{"id": "<account id>", "template": "General List" | "Amortizable" | "Accrual" | "Schedule List"}}, ...]}}
One entry per account. If unsure, choose "General List".
"""


async def _classify_templates(accounts: list) -> dict:
    """Ask Claude to classify each account into a reconciliation template.

    Returns {account_id: template}. Silently drops any entries with invalid
    template names so the caller can fall back to the General List default.
    """
    if not HAS_AGENT_SDK:
        raise RuntimeError("claude-agent-sdk not installed")
    cli_path = _find_claude_cli()
    if not cli_path:
        raise RuntimeError(
            "`claude` CLI not found — install Claude Code and run `claude setup-token`"
        )

    # Build a compact prompt — one line per account, enough context for a solid guess.
    lines = []
    for a in accounts:
        parts = [f"id={a['id']}", f"acct={a['account_number']}"]
        if a.get("description"):
            parts.append(f"desc={a['description']!r}")
        if a.get("account_type"):
            parts.append(f"type={a['account_type']}")
        lines.append("- " + "  ".join(parts))
    prompt = _CLASSIFY_PROMPT_TMPL.format(accounts="\n".join(lines))

    options = ClaudeAgentOptions(
        allowed_tools=[],                # no tools needed — just ask + answer
        permission_mode="bypassPermissions",
        cwd=str(UPLOAD_DIR),
        cli_path=cli_path,
        env={"ANTHROPIC_API_KEY": ""},   # force subscription auth
        max_turns=1,
    )
    if _EXTRACT_MODEL:
        options.model = _EXTRACT_MODEL

    text_out = ""
    async for message in _sdk_query(prompt=prompt, options=options):
        if isinstance(message, ResultMessage):
            if getattr(message, "result", None):
                text_out = message.result
        elif isinstance(message, AssistantMessage):
            for block in getattr(message, "content", []) or []:
                if isinstance(block, TextBlock):
                    text_out += block.text

    raw = _extract_json_object(text_out)
    if not raw:
        raise RuntimeError(f"classifier returned no JSON: {text_out[:300]!r}")
    try:
        data = json.loads(raw)
    except Exception as e:
        raise RuntimeError(f"classifier returned invalid JSON: {e}; raw={raw[:300]!r}")
    items = data.get("classifications") if isinstance(data, dict) else None
    if not isinstance(items, list):
        raise RuntimeError("classifier response missing 'classifications' array")

    out: dict = {}
    for entry in items:
        if not isinstance(entry, dict):
            continue
        aid = str(entry.get("id") or "").strip()
        tmpl = str(entry.get("template") or "").strip()
        if aid and tmpl in _VALID_TEMPLATES:
            out[aid] = tmpl
    return out


def _num(v):
    if v in (None, ""):
        return 0
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0


def _derive_period(row: dict) -> Optional[str]:
    """Try to infer a YYYY-MM period from the columns of a row."""
    # Fiscal_Year + Fiscal_Period
    fy = _find(row, "fiscal_year")
    fp = _find(row, "fiscal_period")
    if fy and fp:
        try:
            return f"{int(fy):04d}-{int(fp):02d}"
        except (ValueError, TypeError):
            pass
    # Period_End_Date
    ped = _find(row, "period_end")
    if ped:
        s = str(ped).strip()
        # ISO date: YYYY-MM-DD...
        m = _re.match(r"^(\d{4})-(\d{1,2})", s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None


@app.get("/api/imports")
def list_imports(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute("SELECT * FROM imports ORDER BY created_at DESC").fetchall()
        return [dict(r) for r in rows]


# ────────────────────────── Reconciliations ──────────────────────────
def _period_end_iso(period: str) -> Optional[str]:
    """'2026-04' → '2026-04-30'. Returns None on bad input."""
    m = _re.match(r"^(\d{4})-(\d{1,2})$", period or "")
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    last_day = monthrange(y, mo)[1]
    return f"{y:04d}-{mo:02d}-{last_day:02d}"


def _effective_items_sum(items, template: str, period: str) -> float:
    """Template-aware supporting-items total.

    For Accrual and Schedule List, only items whose origination date is on or
    before the reconciliation period's last day count — future-scheduled rows
    belong to the forward schedule view, not this period's unidentified
    difference. Every other template uses the naive sum.
    """
    if template not in ("Accrual", "Schedule List"):
        return sum((i["amount"] or 0) for i in items)
    pe = _period_end_iso(period)
    if not pe:
        return sum((i["amount"] or 0) for i in items)
    total = 0.0
    for i in items:
        d = i["origination"] or ""
        # Items without a date count in (treat as current-period).
        if (not d) or d[:10] <= pe:
            total += i["amount"] or 0
    return total


def _serialize_recon(conn, rec):
    row = dict(rec)
    acct = conn.execute("SELECT * FROM accounts WHERE id=?", (row["account_id"],)).fetchone()
    if acct:
        a = dict(acct)
        row["account"] = a["account_number"]
        row["description"] = a["description"]
        row["entity"] = a["entity"]
        row["entity_code"] = a["entity_code"]
        row["template"] = a["template"]
        row["preparer"] = a["preparer"]
        row["approver"] = a["approver"]
        row["currency"] = a["currency"]
        row["cert_threshold_amt"] = a["cert_threshold_amt"]
        row["cert_threshold_pct"] = a["cert_threshold_pct"]
        row["group_id"] = a.get("group_id")
        if a.get("group_id"):
            g = conn.execute(
                "SELECT name FROM account_groups WHERE id=?", (a["group_id"],)
            ).fetchone()
            row["group_name"] = g["name"] if g else None
    # Totals — template-aware. Accrual uses cumulative-through-period so future
    # schedule rows don't affect the current period's unidentified difference.
    items = conn.execute("SELECT * FROM supporting_items WHERE recon_id=?", (row["id"],)).fetchall()
    row["items_total"] = _effective_items_sum(items, row.get("template") or "General List", row["period"])
    row["items_count"] = len(items)
    row["unidentified"] = (row["gl_balance"] or 0) - row["items_total"]
    return row


@app.get("/api/reconciliations")
def list_reconciliations(period: Optional[str] = None, user=Depends(get_user)):
    with tx() as conn:
        if period:
            rows = conn.execute(
                "SELECT * FROM reconciliations WHERE period=? ORDER BY id", (period,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM reconciliations ORDER BY period DESC, id").fetchall()
        result = [_serialize_recon(conn, r) for r in rows]

        # Role-based filtering: Preparer sees own + unassigned; Approver sees own + unassigned;
        # Admin + Auditor see all.
        if user["role"] == "Preparer":
            result = [r for r in result if _visible_to(r, user, "preparer")]
        elif user["role"] == "Approver":
            result = [r for r in result if _visible_to(r, user, "approver")]
        return result


@app.get("/api/reconciliations/{rid}")
def get_recon(rid: str, user=Depends(get_user)):
    with tx() as conn:
        row = conn.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        rec = _serialize_recon(conn, row)
        items = conn.execute(
            "SELECT * FROM supporting_items WHERE recon_id=? ORDER BY created_at",
            (rid,),
        ).fetchall()
        rec["items"] = [dict(i) for i in items]
        # parse extra JSON
        for it in rec["items"]:
            try:
                it["extra"] = json.loads(it["extra"]) if it["extra"] else {}
            except Exception:
                it["extra"] = {}
        comments = conn.execute(
            "SELECT * FROM comments WHERE recon_id=? ORDER BY created_at", (rid,),
        ).fetchall()
        rec["comments"] = [dict(c) for c in comments]
        # Per-recon docs + any group-level docs for this recon's group in this period.
        acct = conn.execute(
            "SELECT group_id FROM accounts WHERE id=?", (row["account_id"],)
        ).fetchone()
        gid = acct["group_id"] if acct else None
        if gid:
            docs = conn.execute(
                """SELECT id, filename, uploaded_by, size_bytes, created_at, group_id
                   FROM documents
                   WHERE recon_id=? OR (group_id=? AND period=?)
                   ORDER BY created_at""",
                (rid, gid, rec["period"]),
            ).fetchall()
        else:
            docs = conn.execute(
                """SELECT id, filename, uploaded_by, size_bytes, created_at, group_id
                   FROM documents WHERE recon_id=? ORDER BY created_at""",
                (rid,),
            ).fetchall()
        rec["documents"] = [dict(d) for d in docs]
        # Attach group info
        if gid:
            g = conn.execute("SELECT * FROM account_groups WHERE id=?", (gid,)).fetchone()
            if g:
                rec["group"] = dict(g)
        return rec


class ItemReq(BaseModel):
    amount: float
    item_class: Optional[str] = None
    origination: Optional[str] = None
    description: Optional[str] = None
    extra: Optional[dict] = None


@app.post("/api/reconciliations/{rid}/items")
def add_item(rid: str, req: ItemReq, user=Depends(require_role("Preparer", "Admin"))):
    iid = f"item-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        rec = conn.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        conn.execute(
            """INSERT INTO supporting_items (id, recon_id, amount, item_class, origination, description, extra)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (iid, rid, req.amount, req.item_class, req.origination, req.description,
             json.dumps(req.extra) if req.extra else None),
        )
        # Moving to "In Progress" if adding an item to a Not Prepared recon
        if rec["status"] == "Not Prepared":
            conn.execute("UPDATE reconciliations SET status='In Progress', updated_at=CURRENT_TIMESTAMP WHERE id=?", (rid,))
    log(user["username"], "add_item", "recon", rid, {"item": iid})
    return {"id": iid}


@app.put("/api/reconciliations/{rid}/items/{iid}")
def update_item(rid: str, iid: str, req: ItemReq, user=Depends(require_role("Preparer", "Admin"))):
    with tx() as conn:
        conn.execute(
            """UPDATE supporting_items SET amount=?, item_class=?, origination=?, description=?, extra=?
               WHERE id=? AND recon_id=?""",
            (req.amount, req.item_class, req.origination, req.description,
             json.dumps(req.extra) if req.extra else None, iid, rid),
        )
    return {"ok": True}


@app.delete("/api/reconciliations/{rid}/items/{iid}")
def delete_item(rid: str, iid: str, user=Depends(require_role("Preparer", "Admin"))):
    with tx() as conn:
        conn.execute("DELETE FROM supporting_items WHERE id=? AND recon_id=?", (iid, rid))
    return {"ok": True}


# ────────────────────────── Invoice PDF extraction (Claude Agent SDK) ──────────────────────────
# Uses the claude-agent-sdk which speaks to the `claude` CLI subprocess.
# Authentication: if the user has run `claude login` / `claude setup-token`,
# the subscription (Max/Pro) is used — no ANTHROPIC_API_KEY required. We
# explicitly blank out ANTHROPIC_API_KEY in the subprocess env so that any
# stale key can't force the SDK back onto API-key auth.
#
# The SDK doesn't accept base64-encoded PDF document blocks the way the raw
# Anthropic API does, so we stage the uploaded PDF on disk and let Claude
# read it via the built-in Read tool (which natively handles PDFs).

_EXTRACT_MODEL = os.environ.get("RECON_EXTRACT_MODEL")  # None → subscription default


_EXTRACT_PROMPT_TMPL = """Read the PDF at {pdf_path} and extract every payable line item.

The PDF contains one or more invoices, bills, receipts, or statements — possibly mixed vendors.
For each DISTINCT payable amount (usually one per invoice; if an invoice has multiple distinct
line items, include each), extract these fields:

- vendor: the issuer name (e.g. "Office Depot", "Verizon", "Blue Cross Blue Shield")
- description: a short human-readable description of what the charge is for
- amount: the charge total as a positive decimal number. No currency symbol, no commas.
  For a multi-line invoice, use the "Total Due", "New Balance", or "Current Charges" value.
- date: the invoice/billing date as YYYY-MM-DD. If only a month is given, use the last day of that month.
- invoice_number: the invoice / transaction / statement number if visible, else an empty string.

Return ONLY a single valid JSON object, no prose, no code fences. Shape:
{{"items": [{{"vendor": "...", "description": "...", "amount": 123.45, "date": "2026-03-29", "invoice_number": "..."}}]}}

If the PDF contains no invoices, return {{"items": []}}.
"""


def _strip_code_fence(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        parts = s.split("\n", 1)
        if len(parts) == 2:
            s = parts[1]
        if s.rstrip().endswith("```"):
            s = s.rstrip()[:-3].rstrip()
    return s


def _extract_json_object(s: str) -> Optional[str]:
    """Pull out the first top-level {...} JSON object from a string, tolerating
    stray prose before/after."""
    s = _strip_code_fence(s)
    # Fast path: already clean JSON
    s_stripped = s.strip()
    if s_stripped.startswith("{") and s_stripped.endswith("}"):
        return s_stripped
    # Slow path: find the first balanced {...} span
    depth = 0
    start = -1
    for i, ch in enumerate(s):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start >= 0:
                return s[start:i + 1]
    return None


@app.post("/api/reconciliations/{rid}/extract-invoice")
async def extract_invoice(rid: str, file: UploadFile = File(...),
                          user=Depends(require_role("Preparer", "Admin"))):
    """Extract line items from an invoice PDF and return them for review.

    This endpoint does NOT create supporting items — the frontend shows the
    results in a review modal and the user picks which ones to commit.
    """
    if not HAS_AGENT_SDK:
        raise HTTPException(
            status_code=500,
            detail="claude-agent-sdk is not installed on the backend. "
                   "Run: pip install claude-agent-sdk>=0.1.0",
        )
    cli_path = _find_claude_cli()
    if not cli_path:
        raise HTTPException(
            status_code=500,
            detail=(
                "The `claude` CLI binary was not found. Install Claude Code "
                "(https://docs.claude.com/claude-code) and sign in with your "
                "Max / Pro subscription via `claude setup-token`. "
                "Then restart the backend."
            ),
        )

    with tx() as conn:
        rec = conn.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Reconciliation not found")

    content = await file.read()
    filename = (file.filename or "invoice.pdf")
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a .pdf")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="PDF exceeds 25 MB limit")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="PDF is empty")

    # Stage the PDF in the uploads directory — the Agent SDK will Read it from
    # there. Nested under a short-lived subdir so we can clean up reliably.
    staging_dir = UPLOAD_DIR / f"extract-{uuid.uuid4().hex[:8]}"
    staging_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = staging_dir / "invoice.pdf"
    pdf_path.write_bytes(content)

    prompt = _EXTRACT_PROMPT_TMPL.format(pdf_path=str(pdf_path))

    # env overrides for the subprocess: blank out ANTHROPIC_API_KEY so the CLI
    # falls back to subscription auth. Everything else inherits from os.environ.
    subprocess_env = {"ANTHROPIC_API_KEY": ""}

    options_kwargs = dict(
        allowed_tools=["Read"],          # ONLY allow reading the PDF
        permission_mode="bypassPermissions",
        cwd=str(staging_dir),
        cli_path=cli_path,
        env=subprocess_env,
        max_turns=3,
    )
    if _EXTRACT_MODEL:
        options_kwargs["model"] = _EXTRACT_MODEL
    options = ClaudeAgentOptions(**options_kwargs)

    text_out = ""
    resolved_model = _EXTRACT_MODEL or "subscription-default"
    try:
        async for message in _sdk_query(prompt=prompt, options=options):
            # ResultMessage carries the final text answer.
            if isinstance(message, ResultMessage):
                if getattr(message, "result", None):
                    text_out = message.result
            # AssistantMessage carries streamed text blocks during the run.
            elif isinstance(message, AssistantMessage):
                for block in getattr(message, "content", []) or []:
                    if isinstance(block, TextBlock):
                        text_out += block.text
    except CLINotFoundError as e:
        raise HTTPException(
            status_code=500,
            detail=(
                f"claude CLI not found: {e}. Install Claude Code "
                "(https://docs.claude.com/claude-code) and sign in."
            ),
        )
    except CLIConnectionError as e:
        raise HTTPException(
            status_code=500,
            detail=(
                f"claude CLI connection failed: {e}. Run `claude setup-token` "
                "to log in with your Max / Pro subscription, then restart the "
                "backend."
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Agent SDK error: {e}")
    finally:
        # Cleanup — leave no PDFs behind.
        try:
            pdf_path.unlink(missing_ok=True)
            staging_dir.rmdir()
        except Exception:
            pass

    raw = _extract_json_object(text_out)
    if not raw:
        raise HTTPException(
            status_code=502,
            detail=f"Claude returned no JSON object. First 500 chars: {text_out[:500]!r}",
        )
    try:
        data = json.loads(raw)
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"Claude returned invalid JSON: {e}. Raw: {raw[:500]}",
        )
    items = data.get("items") if isinstance(data, dict) else None
    if not isinstance(items, list):
        raise HTTPException(status_code=502, detail="Claude response missing 'items' array")

    cleaned = []
    for raw_it in items:
        if not isinstance(raw_it, dict):
            continue
        try:
            amt = float(str(raw_it.get("amount") or 0).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            amt = 0.0
        d = str(raw_it.get("date") or "").strip()[:10]
        m = _re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", d)
        if m:
            d = f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        cleaned.append({
            "vendor":        str(raw_it.get("vendor") or "").strip(),
            "description":   str(raw_it.get("description") or "").strip(),
            "amount":        round(amt, 2),
            "date":          d,
            "invoice_number": str(raw_it.get("invoice_number") or "").strip(),
        })

    log(user["username"], "extract_invoice", "recon", rid,
        {"filename": filename, "count": len(cleaned), "model": resolved_model,
         "auth": "subscription"})

    return {
        "items": cleaned,
        "count": len(cleaned),
        "model": resolved_model,
        "auth": "claude-code-subscription",
        "source_filename": filename,
    }


class CommentReq(BaseModel):
    text: str


@app.post("/api/reconciliations/{rid}/comments")
def add_comment(rid: str, req: CommentReq, user=Depends(get_user)):
    cid = f"c-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        conn.execute(
            "INSERT INTO comments (id, recon_id, author, text) VALUES (?, ?, ?, ?)",
            (cid, rid, user["name"], req.text),
        )
    return {"id": cid}


@app.post("/api/reconciliations/{rid}/certify")
def certify(rid: str, user=Depends(require_role("Preparer", "Admin"))):
    with tx() as conn:
        rec = conn.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Not found")
        items = conn.execute("SELECT * FROM supporting_items WHERE recon_id=?", (rid,)).fetchall()
        acct = conn.execute("SELECT * FROM accounts WHERE id=?", (rec["account_id"],)).fetchone()
        tpl = acct["template"] if acct else "General List"
        total = _effective_items_sum(items, tpl, rec["period"])
        unidentified = (rec["gl_balance"] or 0) - total
        thresh_amt = (acct["cert_threshold_amt"] or 0) if acct else 0
        thresh_pct = (acct["cert_threshold_pct"] or 0) if acct else 0
        pct_allowed = abs(rec["gl_balance"] or 0) * thresh_pct / 100.0 if thresh_pct else 0
        tolerance = max(thresh_amt, pct_allowed, 0.01)
        if abs(unidentified) > tolerance:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot certify. Unidentified difference {unidentified:.2f} exceeds tolerance {tolerance:.2f}",
            )
        conn.execute(
            """UPDATE reconciliations SET status='Pending Approval', prep_date=?,
               certified_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (datetime.now().strftime("%m/%d/%Y"), user["name"], rid),
        )
    log(user["username"], "certify", "recon", rid)
    return {"ok": True, "status": "Pending Approval"}


@app.post("/api/reconciliations/{rid}/approve")
def approve(rid: str, user=Depends(require_role("Approver", "Admin"))):
    with tx() as conn:
        conn.execute(
            """UPDATE reconciliations SET status='Reviewed', app_date=?, approved_by=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (datetime.now().strftime("%m/%d/%Y"), user["name"], rid),
        )
    log(user["username"], "approve", "recon", rid)
    return {"ok": True, "status": "Reviewed"}


@app.post("/api/reconciliations/{rid}/reject")
def reject(rid: str, req: RejectReq, user=Depends(require_role("Approver", "Admin"))):
    with tx() as conn:
        conn.execute(
            """UPDATE reconciliations SET status='In Progress', reject_reason=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (req.reason, rid),
        )
        if req.reason:
            conn.execute(
                "INSERT INTO comments (id, recon_id, author, text) VALUES (?, ?, ?, ?)",
                (f"c-{uuid.uuid4().hex[:8]}", rid, user["name"], f"[REJECTED] {req.reason}"),
            )
    log(user["username"], "reject", "recon", rid, {"reason": req.reason})
    return {"ok": True, "status": "In Progress"}


# ────────────────────────── Supporting Documents ──────────────────────────
@app.post("/api/reconciliations/{rid}/documents")
async def upload_doc(rid: str, file: UploadFile = File(...), user=Depends(get_user)):
    """Upload a supporting document for a reconciliation.

    If the reconciliation's account belongs to a group, the document is stored
    as a group-level document for that period — visible to every recon in the
    group. A single upload therefore satisfies the whole group.
    """
    content = await file.read()
    did = f"doc-{uuid.uuid4().hex[:8]}"
    fname = file.filename or "unnamed"
    # Sanitise filename — FileResponse needs the stored path to be locatable.
    safe_name = _re.sub(r"[^A-Za-z0-9._-]+", "_", fname).strip("_") or "doc"
    save_path = UPLOAD_DIR / f"{did}__{safe_name}"
    save_path.write_bytes(content)
    with tx() as conn:
        rec = conn.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        acct = conn.execute(
            "SELECT * FROM accounts WHERE id=?", (rec["account_id"],)
        ).fetchone()
        gid = acct["group_id"] if acct else None
        if gid:
            conn.execute(
                """INSERT INTO documents (id, recon_id, group_id, period, filename,
                   stored_path, uploaded_by, size_bytes)
                   VALUES (?, NULL, ?, ?, ?, ?, ?, ?)""",
                (did, gid, rec["period"], fname, str(save_path),
                 user["name"], len(content)),
            )
        else:
            conn.execute(
                """INSERT INTO documents (id, recon_id, filename, stored_path,
                   uploaded_by, size_bytes)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (did, rid, fname, str(save_path), user["name"], len(content)),
            )
    return {"id": did, "filename": fname, "group_shared": bool(gid)}


@app.get("/api/documents/{did}")
def get_doc(did: str, user=Depends(get_user)):
    with tx() as conn:
        d = conn.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone()
        if not d:
            raise HTTPException(status_code=404, detail="Not found")
        import mimetypes
        p = Path(d["stored_path"])
        if not p.exists():
            raise HTTPException(status_code=404, detail="File missing on disk")
        ctype, _ = mimetypes.guess_type(d["filename"])
        return FileResponse(str(p), filename=d["filename"],
                            media_type=ctype or "application/octet-stream")


@app.delete("/api/documents/{did}")
def delete_doc(did: str, user=Depends(get_user)):
    with tx() as conn:
        d = conn.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone()
        if not d:
            raise HTTPException(status_code=404, detail="Not found")
        # Only the uploader or an Admin can delete.
        if user["role"] != "Admin" and d["uploaded_by"] != user["name"]:
            raise HTTPException(status_code=403, detail="Can only delete your own uploads")
        try:
            Path(d["stored_path"]).unlink()
        except Exception:
            pass
        conn.execute("DELETE FROM documents WHERE id=?", (did,))
    return {"ok": True}


# ────────────────────────── Summary (for Reconciliation Summary page) ──────────────────────────
@app.get("/api/summary")
def summary(period: Optional[str] = None, user=Depends(get_user)):
    with tx() as conn:
        if period:
            rows = conn.execute(
                "SELECT * FROM reconciliations WHERE period=?", (period,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM reconciliations").fetchall()
        recs = [_serialize_recon(conn, r) for r in rows]

        # Role filter (unassigned recons are visible to all Preparers/Approvers as a pool).
        if user["role"] == "Preparer":
            recs = [r for r in recs if _visible_to(r, user, "preparer")]
        elif user["role"] == "Approver":
            recs = [r for r in recs if _visible_to(r, user, "approver")]

        total = len(recs)
        completed_statuses = {"Reviewed", "Approved", "System Certified"}
        pending_statuses   = {"Not Prepared", "In Progress", "Pending Approval"}

        completed     = sum(1 for r in recs if r["status"] in completed_statuses)
        not_completed = sum(1 for r in recs if r["status"] in pending_statuses)

        # Granular counts (useful for future detail)
        by_status = {}
        for r in recs:
            by_status[r["status"]] = by_status.get(r["status"], 0) + 1

        return {
            "period": period,
            "total": total,
            "completed": completed,
            "not_completed": not_completed,
            "by_status": by_status,
        }


# ────────────────────────── Periods ──────────────────────────
_VALID_PERIOD_STATUSES = ("Future", "Open", "Soft-Close", "Closed", "Reopened")


def _default_period_status(period: str) -> str:
    """Infer a sensible default based on where the period sits relative to
    the current calendar month — previous = Closed, current = Open, future =
    Future. Used when a period hasn't been explicitly set yet."""
    if not _re.match(r"^\d{4}-\d{2}$", period or ""):
        return "Open"
    now = datetime.now()
    cur = f"{now.year:04d}-{now.month:02d}"
    if period < cur:
        return "Closed"
    if period == cur:
        return "Open"
    return "Future"


@app.get("/api/periods")
def periods(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute(
            "SELECT DISTINCT period FROM reconciliations ORDER BY period DESC"
        ).fetchall()
        return [r["period"] for r in rows]


@app.get("/api/period-statuses")
def list_period_statuses(user=Depends(get_user)):
    """Return the status for every period we know about (from recons and from
    previously-changed statuses), plus the currently-selected calendar month.
    Anything without an explicit status falls back to the default."""
    with tx() as conn:
        seen = set()
        for r in conn.execute("SELECT DISTINCT period FROM reconciliations").fetchall():
            seen.add(r["period"])
        for r in conn.execute("SELECT period FROM period_statuses").fetchall():
            seen.add(r["period"])
        now = datetime.now()
        seen.add(f"{now.year:04d}-{now.month:02d}")

        rows_by_period = {
            r["period"]: dict(r)
            for r in conn.execute("SELECT * FROM period_statuses").fetchall()
        }

        out = []
        for p in sorted(seen, reverse=True):
            explicit = rows_by_period.get(p)
            if explicit:
                out.append({
                    "period":     p,
                    "status":     explicit["status"],
                    "changed_by": explicit.get("changed_by"),
                    "changed_at": explicit.get("changed_at"),
                    "explicit":   True,
                })
            else:
                out.append({
                    "period":     p,
                    "status":     _default_period_status(p),
                    "changed_by": None,
                    "changed_at": None,
                    "explicit":   False,
                })
        return out


class PeriodStatusReq(BaseModel):
    status: str


@app.put("/api/period-statuses/{period}")
def set_period_status(period: str, req: PeriodStatusReq,
                      user=Depends(require_role("Admin"))):
    if req.status not in _VALID_PERIOD_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Expected one of {', '.join(_VALID_PERIOD_STATUSES)}",
        )
    if not _re.match(r"^\d{4}-\d{2}$", period or ""):
        raise HTTPException(status_code=400, detail="Period must be YYYY-MM")
    with tx() as conn:
        existing = conn.execute(
            "SELECT period FROM period_statuses WHERE period=?", (period,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE period_statuses SET status=?, changed_by=?, "
                "changed_at=CURRENT_TIMESTAMP WHERE period=?",
                (req.status, user["username"], period),
            )
        else:
            conn.execute(
                "INSERT INTO period_statuses (period, status, changed_by) VALUES (?, ?, ?)",
                (period, req.status, user["username"]),
            )
    log(user["username"], "set_period_status", "period", period, {"status": req.status})
    return {"ok": True, "period": period, "status": req.status}


# ────────────────────────── Auto-certification rules ──────────────────────────
def _prev_period(period: str) -> Optional[str]:
    """'2026-04' → '2026-03'. Returns None if the input is malformed."""
    m = _re.match(r"^(\d{4})-(\d{1,2})$", period or "")
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    mo -= 1
    if mo == 0:
        mo = 12
        y -= 1
    return f"{y:04d}-{mo:02d}"


def _tolerance_for(acct: dict, gl_balance: float) -> float:
    pct = float(acct.get("cert_threshold_pct") or 0)
    amt = float(acct.get("cert_threshold_amt") or 0)
    pct_amt = abs(gl_balance or 0) * pct / 100.0 if pct else 0
    # Never below one cent — avoids floating-point false negatives.
    return max(amt, pct_amt, 0.01)


# ────────────────────────── Auto-recon rules CRUD ──────────────────────────
@app.get("/api/auto-rules")
def list_auto_rules(user=Depends(get_user)):
    with tx() as conn:
        rows = conn.execute(
            "SELECT id, name, description, enabled, updated_at FROM auto_recon_rules ORDER BY id"
        ).fetchall()
        return [
            {**dict(r), "enabled": bool(r["enabled"])}
            for r in rows
        ]


class AutoRuleReq(BaseModel):
    enabled: bool


@app.put("/api/auto-rules/{rule_id}")
def update_auto_rule(rule_id: str, req: AutoRuleReq,
                     user=Depends(require_role("Admin"))):
    if rule_id not in ("rule1", "rule2", "rule3"):
        raise HTTPException(status_code=404, detail="Unknown rule")
    with tx() as conn:
        conn.execute(
            "UPDATE auto_recon_rules SET enabled=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (1 if req.enabled else 0, rule_id),
        )
    log(user["username"], "toggle_rule", "rule", rule_id, {"enabled": req.enabled})
    return {"ok": True, "id": rule_id, "enabled": req.enabled}


def _enabled_rules(conn) -> set:
    rows = conn.execute(
        "SELECT id FROM auto_recon_rules WHERE enabled=1"
    ).fetchall()
    return {r["id"] for r in rows}


@app.post("/api/auto-certify")
def auto_certify(period: str, user=Depends(require_role("Admin"))):
    """Evaluate the 3 PRD auto-reconciliation rules against the given period
    and transition matching reconciliations to 'System Certified'. Only
    rules that are toggled on in the Auto-recon rules page will fire."""
    if not period:
        raise HTTPException(status_code=400, detail="period is required (YYYY-MM)")
    prev = _prev_period(period)
    certified = []
    by_rule = {"rule1": 0, "rule2": 0, "rule3": 0}
    skipped = []

    with tx() as conn:
        enabled = _enabled_rules(conn)
        if not enabled:
            return {
                "period": period,
                "certified_count": 0,
                "by_rule": by_rule,
                "certified": [],
                "skipped_count": 0,
                "warning": "No auto-recon rules are enabled. Enable at least one on the "
                           "Auto-recon rules page.",
            }

        recs = conn.execute(
            "SELECT * FROM reconciliations WHERE period=?", (period,)
        ).fetchall()

        for r in recs:
            r = dict(r)
            # Already done — leave it alone.
            if r["status"] in ("Reviewed", "Approved", "System Certified"):
                continue

            acct_row = conn.execute(
                "SELECT * FROM accounts WHERE id=?", (r["account_id"],)
            ).fetchone()
            if not acct_row:
                continue
            acct = dict(acct_row)

            items = conn.execute(
                "SELECT * FROM supporting_items WHERE recon_id=?", (r["id"],)
            ).fetchall()
            items_count = len(items)
            items_total = _effective_items_sum(items, acct["template"], r["period"])
            gl = float(r["gl_balance"] or 0)
            tol = _tolerance_for(acct, gl)

            fired = None

            # Rule 1: Zero Balance, No Activity
            if "rule1" in enabled and abs(gl) < 0.005 and items_count == 0:
                fired = "rule1"

            # Rule 2: Schedule Match (only for schedule-backed templates)
            elif "rule2" in enabled and acct["template"] in ("Amortizable", "Accrual") and items_count > 0:
                if abs(gl - items_total) <= tol:
                    fired = "rule2"

            # Rule 3: Balance Unchanged from Prior Period
            elif "rule3" in enabled and prev:
                prior = conn.execute(
                    "SELECT * FROM reconciliations WHERE account_id=? AND period=?",
                    (r["account_id"], prev),
                ).fetchone()
                if prior:
                    prior = dict(prior)
                    prior_done = prior["status"] in ("Reviewed", "Approved", "System Certified")
                    if (
                        prior_done
                        and abs(gl - float(prior["gl_balance"] or 0)) < 0.005
                        and items_count == 0
                    ):
                        fired = "rule3"

            if not fired:
                skipped.append({"id": r["id"], "account": acct["account_number"], "reason": "no rule matched"})
                continue

            conn.execute(
                """UPDATE reconciliations SET status='System Certified',
                   certified_by='System', prep_date=?, updated_at=CURRENT_TIMESTAMP
                   WHERE id=?""",
                (datetime.now().strftime("%m/%d/%Y"), r["id"]),
            )
            conn.execute(
                "INSERT INTO comments (id, recon_id, author, text) VALUES (?, ?, ?, ?)",
                (
                    f"c-{uuid.uuid4().hex[:8]}",
                    r["id"],
                    "System",
                    {
                        "rule1": "Auto-certified: zero balance with no activity.",
                        "rule2": "Auto-certified: schedule balance matches GL balance within tolerance.",
                        "rule3": "Auto-certified: balance unchanged from prior period.",
                    }[fired],
                ),
            )
            by_rule[fired] += 1
            certified.append({"id": r["id"], "account": acct["account_number"], "rule": fired})

    log(user["username"], "auto_certify", "period", period, {"certified": len(certified), "by_rule": by_rule})
    return {
        "period": period,
        "certified_count": len(certified),
        "by_rule": by_rule,
        "certified": certified,
        "skipped_count": len(skipped),
    }


# ────────────────────────── Reset (demo helper) ──────────────────────────
@app.post("/api/reset")
def reset(user=Depends(require_role("Admin"))):
    with tx() as conn:
        for t in ("supporting_items", "comments", "documents", "imports",
                  "reconciliations", "accounts", "account_groups", "audit_log",
                  "data_source_runs"):
            conn.execute(f"DELETE FROM {t}")
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# Data sources — scheduled trial-balance ingestion from Local / SFTP / S3
# ═══════════════════════════════════════════════════════════════════════════════

import threading
import fnmatch
import glob as _glob
import time as _time

# Optional heavy deps — import lazily so missing them doesn't break the app.
try:
    import paramiko  # SFTP
    HAS_PARAMIKO = True
except ImportError:
    HAS_PARAMIKO = False

try:
    import boto3     # S3
    HAS_BOTO3 = True
except ImportError:
    HAS_BOTO3 = False


_VALID_SOURCE_TYPES = ("local", "sftp", "s3")


class DataSourceReq(BaseModel):
    name: str
    type: str
    config: dict
    file_pattern: Optional[str] = "*.xlsx"
    period_rule: Optional[str] = "mtime"           # 'mtime' | 'filename' | 'current-month'
    period_regex: Optional[str] = None              # only used when period_rule='filename'
    auto_classify: bool = False
    schedule_minutes: Optional[int] = None
    enabled: bool = True


def _serialize_source(row) -> dict:
    d = dict(row)
    try:
        d["config"] = json.loads(d["config"]) if d.get("config") else {}
    except Exception:
        d["config"] = {}
    # Scrub secrets from the serialized form — UI only needs to know they exist.
    redacted = dict(d["config"])
    for k in ("password", "secret_access_key", "private_key"):
        if k in redacted and redacted[k]:
            redacted[k] = "••••"
    d["config"] = redacted
    d["enabled"] = bool(d.get("enabled"))
    d["auto_classify"] = bool(d.get("auto_classify"))
    # Compute next-run-at
    if d.get("schedule_minutes") and d.get("last_run_at"):
        try:
            last = datetime.strptime(d["last_run_at"], "%Y-%m-%d %H:%M:%S")
            d["next_run_at"] = (last + timedelta(minutes=int(d["schedule_minutes"]))).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            d["next_run_at"] = None
    elif d.get("schedule_minutes") and not d.get("last_run_at"):
        d["next_run_at"] = "pending first run"
    else:
        d["next_run_at"] = None
    return d


@app.get("/api/data-sources")
def list_data_sources(user=Depends(require_role("Admin"))):
    with tx() as conn:
        rows = conn.execute("SELECT * FROM data_sources ORDER BY created_at DESC").fetchall()
        return [_serialize_source(r) for r in rows]


@app.get("/api/data-sources/{sid}")
def get_data_source(sid: str, user=Depends(require_role("Admin"))):
    with tx() as conn:
        r = conn.execute("SELECT * FROM data_sources WHERE id=?", (sid,)).fetchone()
        if not r:
            raise HTTPException(status_code=404, detail="Data source not found")
        out = _serialize_source(r)
        runs = conn.execute(
            "SELECT * FROM data_source_runs WHERE source_id=? ORDER BY started_at DESC LIMIT 20",
            (sid,),
        ).fetchall()
        out["recent_runs"] = [dict(x) for x in runs]
        return out


@app.post("/api/data-sources")
def create_data_source(req: DataSourceReq, user=Depends(require_role("Admin"))):
    _validate_source_req(req)
    sid = f"src-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        conn.execute(
            """INSERT INTO data_sources
               (id, name, type, config, file_pattern, period_rule, period_regex,
                auto_classify, schedule_minutes, enabled, created_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (sid, req.name.strip(), req.type, json.dumps(req.config or {}),
             req.file_pattern or "*.xlsx", req.period_rule or "mtime",
             req.period_regex, int(req.auto_classify),
             req.schedule_minutes, int(req.enabled), user["username"]),
        )
    log(user["username"], "create_data_source", "source", sid, {"type": req.type, "name": req.name})
    return {"id": sid}


@app.put("/api/data-sources/{sid}")
def update_data_source(sid: str, req: DataSourceReq, user=Depends(require_role("Admin"))):
    _validate_source_req(req)
    with tx() as conn:
        existing = conn.execute("SELECT * FROM data_sources WHERE id=?", (sid,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Data source not found")
        # Preserve any secret fields the UI redacted with "••••" on read.
        merged_config = dict(req.config or {})
        try:
            old_config = json.loads(existing["config"]) if existing["config"] else {}
        except Exception:
            old_config = {}
        for secret_key in ("password", "secret_access_key", "private_key"):
            if merged_config.get(secret_key) == "••••" and secret_key in old_config:
                merged_config[secret_key] = old_config[secret_key]
        conn.execute(
            """UPDATE data_sources SET name=?, type=?, config=?, file_pattern=?,
               period_rule=?, period_regex=?, auto_classify=?, schedule_minutes=?,
               enabled=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (req.name.strip(), req.type, json.dumps(merged_config),
             req.file_pattern or "*.xlsx", req.period_rule or "mtime",
             req.period_regex, int(req.auto_classify),
             req.schedule_minutes, int(req.enabled), sid),
        )
    return {"ok": True}


@app.delete("/api/data-sources/{sid}")
def delete_data_source(sid: str, user=Depends(require_role("Admin"))):
    with tx() as conn:
        conn.execute("DELETE FROM data_sources WHERE id=?", (sid,))
    log(user["username"], "delete_data_source", "source", sid)
    return {"ok": True}


@app.post("/api/data-sources/{sid}/test")
async def test_data_source(sid: str, user=Depends(require_role("Admin"))):
    """Connect to the source and list files matching the pattern — without
    ingesting anything. Used by the setup wizard's 'Test connection' button."""
    with tx() as conn:
        row = conn.execute("SELECT * FROM data_sources WHERE id=?", (sid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Data source not found")
        src = dict(row)
        try:
            src["config"] = json.loads(src["config"]) if src["config"] else {}
        except Exception:
            src["config"] = {}
    try:
        handler = _get_handler(src["type"])
        files = handler.list_files(src["config"], src.get("file_pattern") or "*.xlsx")
        return {
            "ok": True,
            "type": src["type"],
            "files_found": len(files),
            "sample": [f["name"] for f in files[:10]],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/data-sources/{sid}/run")
async def run_data_source_now(sid: str, user=Depends(require_role("Admin"))):
    with tx() as conn:
        row = conn.execute("SELECT * FROM data_sources WHERE id=?", (sid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Data source not found")
    res = await _run_data_source(sid, triggered_by=user["username"])
    return res


def _validate_source_req(req: DataSourceReq):
    if req.type not in _VALID_SOURCE_TYPES:
        raise HTTPException(status_code=400, detail=f"type must be one of {', '.join(_VALID_SOURCE_TYPES)}")
    if req.period_rule and req.period_rule not in ("mtime", "filename", "current-month"):
        raise HTTPException(status_code=400, detail="period_rule must be mtime|filename|current-month")
    if req.schedule_minutes is not None and req.schedule_minutes < 0:
        raise HTTPException(status_code=400, detail="schedule_minutes must be positive")


# ───────── source handlers ─────────
class LocalSource:
    type = "local"

    def list_files(self, cfg, pattern):
        folder = Path(cfg.get("folder_path") or "").expanduser()
        if not folder.is_dir():
            raise RuntimeError(f"Folder does not exist: {folder}")
        out = []
        for p in folder.glob(pattern):
            if p.is_file():
                out.append({
                    "name": p.name,
                    "path": str(p),
                    "size": p.stat().st_size,
                    "mtime": p.stat().st_mtime,
                })
        out.sort(key=lambda f: f["mtime"], reverse=True)
        return out

    def fetch(self, cfg, file):
        return Path(file["path"]).read_bytes()


class SFTPSource:
    type = "sftp"

    def _client(self, cfg):
        if not HAS_PARAMIKO:
            raise RuntimeError(
                "paramiko is not installed on the backend. Run: pip install paramiko"
            )
        host = cfg.get("host")
        port = int(cfg.get("port") or 22)
        username = cfg.get("username")
        password = cfg.get("password")
        if not host or not username:
            raise RuntimeError("host and username are required")
        transport = paramiko.Transport((host, port))
        try:
            if password:
                transport.connect(username=username, password=password)
            else:
                # Could extend to SSH-key auth via cfg['private_key']
                raise RuntimeError("Only password auth is supported in this prototype")
            return paramiko.SFTPClient.from_transport(transport), transport
        except Exception:
            transport.close()
            raise

    def list_files(self, cfg, pattern):
        sftp, transport = self._client(cfg)
        try:
            remote_dir = cfg.get("remote_path") or "."
            out = []
            for entry in sftp.listdir_attr(remote_dir):
                if fnmatch.fnmatch(entry.filename, pattern):
                    out.append({
                        "name": entry.filename,
                        "path": f"{remote_dir.rstrip('/')}/{entry.filename}",
                        "size": entry.st_size,
                        "mtime": entry.st_mtime,
                    })
            out.sort(key=lambda f: f["mtime"], reverse=True)
            return out
        finally:
            try: sftp.close()
            except Exception: pass
            try: transport.close()
            except Exception: pass

    def fetch(self, cfg, file):
        sftp, transport = self._client(cfg)
        try:
            with sftp.open(file["path"], "rb") as f:
                return f.read()
        finally:
            try: sftp.close()
            except Exception: pass
            try: transport.close()
            except Exception: pass


class S3Source:
    type = "s3"

    def _client(self, cfg):
        if not HAS_BOTO3:
            raise RuntimeError(
                "boto3 is not installed on the backend. Run: pip install boto3"
            )
        return boto3.client(
            "s3",
            region_name=cfg.get("region") or "us-east-1",
            aws_access_key_id=cfg.get("access_key_id") or None,
            aws_secret_access_key=cfg.get("secret_access_key") or None,
        )

    def list_files(self, cfg, pattern):
        bucket = cfg.get("bucket")
        prefix = cfg.get("prefix") or ""
        if not bucket:
            raise RuntimeError("bucket is required")
        c = self._client(cfg)
        out = []
        resp = c.list_objects_v2(Bucket=bucket, Prefix=prefix)
        for obj in resp.get("Contents", []):
            key = obj["Key"]
            name = key.rsplit("/", 1)[-1]
            if not name or not fnmatch.fnmatch(name, pattern):
                continue
            out.append({
                "name": name,
                "path": key,
                "size": obj.get("Size", 0),
                "mtime": obj["LastModified"].timestamp(),
            })
        out.sort(key=lambda f: f["mtime"], reverse=True)
        return out

    def fetch(self, cfg, file):
        bucket = cfg.get("bucket")
        c = self._client(cfg)
        obj = c.get_object(Bucket=bucket, Key=file["path"])
        return obj["Body"].read()


def _get_handler(type_: str):
    if type_ == "local": return LocalSource()
    if type_ == "sftp":  return SFTPSource()
    if type_ == "s3":    return S3Source()
    raise RuntimeError(f"Unknown source type: {type_}")


def _derive_period_for_file(src: dict, file: dict) -> Optional[str]:
    rule = src.get("period_rule") or "mtime"
    if rule == "current-month":
        now = datetime.now()
        return f"{now.year:04d}-{now.month:02d}"
    if rule == "filename":
        rx = src.get("period_regex") or ""
        if not rx:
            return None
        m = _re.search(rx, file["name"])
        if not m:
            return None
        # Expected 1 or 2 groups: YYYY (and optional MM). Else try the whole match.
        gs = m.groups()
        if len(gs) >= 2 and gs[0] and gs[1]:
            try:
                return f"{int(gs[0]):04d}-{int(gs[1]):02d}"
            except (ValueError, TypeError):
                pass
        # Fall back: whole match as YYYY-MM
        if _re.match(r"^\d{4}-\d{2}$", m.group(0)):
            return m.group(0)
        return None
    # mtime: the file's mtime (assumed uploaded at month-end).
    try:
        d = datetime.fromtimestamp(file["mtime"])
        return f"{d.year:04d}-{d.month:02d}"
    except Exception:
        return None


async def _run_data_source(sid: str, *, triggered_by: str) -> dict:
    """Execute a single source run. Records a row in data_source_runs and
    updates the source's last_run_at / last_status."""
    run_id = f"run-{uuid.uuid4().hex[:8]}"
    with tx() as conn:
        row = conn.execute("SELECT * FROM data_sources WHERE id=?", (sid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Data source not found")
        src = dict(row)
    try:
        src["config"] = json.loads(src["config"]) if src["config"] else {}
    except Exception:
        src["config"] = {}

    with tx() as conn:
        conn.execute(
            "INSERT INTO data_source_runs (id, source_id, triggered_by, status) VALUES (?, ?, ?, 'running')",
            (run_id, sid, triggered_by),
        )

    summary = {
        "files_processed": 0,
        "accounts_created": 0,
        "accounts_updated": 0,
        "per_file": [],
    }
    status = "ok"
    error_msg = None
    cutoff = float(src.get("last_mtime") or 0)
    max_mtime_seen = cutoff

    try:
        handler = _get_handler(src["type"])
        files = handler.list_files(src["config"], src.get("file_pattern") or "*.xlsx")

        # Only ingest files with a strictly greater mtime than any file we've
        # already pulled from this source. Using a stored float last_mtime
        # (sub-second precision) avoids re-ingesting a file on the next run.
        new_files = [f for f in files if f.get("mtime", 0) > cutoff]

        if not new_files:
            status = "no-new-files"
        else:
            for f in new_files:
                entry = {"name": f["name"], "status": "ok"}
                try:
                    content = handler.fetch(src["config"], f)
                    period = _derive_period_for_file(src, f)
                    res = await _ingest_trial_balance(
                        content, f["name"],
                        period=period,
                        classify=bool(src.get("auto_classify")),
                        username=f"scheduler:{src['name']}",
                    )
                    entry.update({
                        "period": res["period"],
                        "accounts_created": res["accounts_created"],
                        "accounts_updated": res["accounts_updated"],
                    })
                    summary["files_processed"] += 1
                    summary["accounts_created"] += res["accounts_created"]
                    summary["accounts_updated"] += res["accounts_updated"]
                    # Track the highest mtime we successfully ingested.
                    if f.get("mtime", 0) > max_mtime_seen:
                        max_mtime_seen = f["mtime"]
                except Exception as e:
                    entry["status"] = "error"
                    entry["error"] = str(e)
                    status = "error"
                    error_msg = error_msg or f"{f['name']}: {e}"
                summary["per_file"].append(entry)
    except Exception as e:
        status = "error"
        error_msg = str(e)

    with tx() as conn:
        conn.execute(
            """UPDATE data_source_runs SET ended_at=CURRENT_TIMESTAMP,
                  status=?, files_processed=?, accounts_created=?,
                  accounts_updated=?, error=?, details=? WHERE id=?""",
            (status, summary["files_processed"], summary["accounts_created"],
             summary["accounts_updated"], error_msg, json.dumps(summary), run_id),
        )
        # Only bump last_mtime when we actually ingested something — this way
        # a failed/empty run doesn't accidentally skip files on the next pass.
        if max_mtime_seen > cutoff:
            conn.execute(
                """UPDATE data_sources SET last_run_at=CURRENT_TIMESTAMP,
                      last_mtime=?, last_status=?, last_error=?,
                      updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (max_mtime_seen, status, error_msg, sid),
            )
        else:
            conn.execute(
                """UPDATE data_sources SET last_run_at=CURRENT_TIMESTAMP,
                      last_status=?, last_error=?,
                      updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (status, error_msg, sid),
            )

    log(triggered_by, "run_data_source", "source", sid,
        {"run_id": run_id, "status": status, "summary": summary})

    return {"run_id": run_id, "status": status, "error": error_msg, **summary}


# ───────── background scheduler ─────────
_scheduler_thread = None
_scheduler_stop = threading.Event()


def _scheduler_loop():
    """Wakes up every 30 seconds, inspects enabled data sources with a
    schedule, fires any whose next_run_at <= now. Single-threaded —
    concurrent runs of the same source aren't a concern for this prototype."""
    import asyncio
    while not _scheduler_stop.is_set():
        try:
            due = []
            with tx() as conn:
                rows = conn.execute(
                    "SELECT * FROM data_sources WHERE enabled=1 AND schedule_minutes IS NOT NULL"
                ).fetchall()
                now = datetime.now()
                for r in rows:
                    r = dict(r)
                    interval = int(r["schedule_minutes"] or 0)
                    if interval <= 0:
                        continue
                    if not r.get("last_run_at"):
                        due.append(r)
                        continue
                    try:
                        last = datetime.strptime(r["last_run_at"], "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        due.append(r)
                        continue
                    if (now - last).total_seconds() >= interval * 60:
                        due.append(r)
            for r in due:
                try:
                    asyncio.run(_run_data_source(r["id"], triggered_by="scheduler"))
                except Exception as e:
                    print(f"[scheduler] {r['id']} {r['name']} failed: {e}")
        except Exception as e:
            print(f"[scheduler] loop error: {e}")
        _scheduler_stop.wait(30)


@app.on_event("startup")
def _start_scheduler():
    global _scheduler_thread
    if _scheduler_thread and _scheduler_thread.is_alive():
        return
    _scheduler_stop.clear()
    _scheduler_thread = threading.Thread(target=_scheduler_loop, daemon=True, name="recon-scheduler")
    _scheduler_thread.start()
    print("[scheduler] started")


@app.on_event("shutdown")
def _stop_scheduler():
    _scheduler_stop.set()


@app.get("/")
def root():
    return {"name": "RECON API", "status": "ok"}
